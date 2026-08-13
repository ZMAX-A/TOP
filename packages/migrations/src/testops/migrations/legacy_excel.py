"""Import the legacy Excel case catalog into an immutable case baseline.

The converter intentionally mirrors the legacy StepExecutor's positional
rules while making every implicit choice explicit in the generated contract.
The source workbook is read only and is never copied into this repository.
"""

from __future__ import annotations

import hashlib
import math
from collections import Counter
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import UUID, uuid5

from pydantic import ValidationError

from testops.contracts import (
    AssertionDefinition,
    CaseBaseline,
    CaseBaselineSource,
    CaseDefinition,
    CasePriority,
    CaseSourceTrace,
    StepDefinition,
    canonical_json_bytes,
    get_assertion_spec,
    get_operation_spec,
)

DEFAULT_WORKSHEET = "自动化测试用例"
BASELINE_NAMESPACE = UUID("a89b783b-1c2e-5a2e-8f3e-e877142d7b4f")

REQUIRED_COLUMNS = (
    "用例ID",
    "模块",
    "测试场景",
    "测试点",
    "优先级",
    "前置条件",
    "操作步骤",
    "元素定位器",
    "操作类型",
    "输入数据",
    "数据类型",
    "期望结果",
    "验证点",
    "断言类型",
    "超时(秒)",
    "是否执行",
    "备注",
)

MODULE_KEYS: Mapping[str, str] = {
    "账号登录": "login",
    "首页": "home",
    "首页搜索": "home-search",
    "首页跳转": "home-navigation",
    "顾客列表": "customer-list",
    "顾客详情": "customer-detail",
    "影像阅览": "image-viewer",
    "案例库": "case-library",
    "个人中心": "profile",
}

_PRIORITY_ALIASES: Mapping[str, CasePriority] = {
    "P0": CasePriority.P0,
    "P1": CasePriority.P1,
    "P2": CasePriority.P2,
    "P3": CasePriority.P3,
    "中": CasePriority.P2,
}
_ENABLED_VALUES = {"是", "1", "Y", "TRUE"}
_DISABLED_VALUES = {"否", "0", "N", "FALSE"}


class LegacyExcelMigrationError(ValueError):
    """Raised when the source cannot be converted without losing semantics."""


@dataclass(frozen=True, slots=True)
class MigrationResult:
    baseline: CaseBaseline
    audit: dict[str, Any]


def _digest_bytes(payload: bytes) -> str:
    return f"sha256:{hashlib.sha256(payload).hexdigest()}"


def _cell_text(value: object, *, strip: bool = True) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        text = "true" if value else "false"
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if not math.isfinite(value):
            raise LegacyExcelMigrationError("Excel contains a non-finite numeric value")
        text = str(int(value)) if value.is_integer() else format(value, ".15g")
    elif isinstance(value, (datetime, date)):
        text = value.isoformat()
    else:
        text = str(value)
    return text.strip() if strip else text


def _required_text(row: Mapping[str, object], column: str) -> str:
    value = _cell_text(row.get(column))
    if not value:
        raise LegacyExcelMigrationError(f"{column}为空")
    return value


def _split_legacy_data(value: object) -> tuple[list[str], str | None]:
    raw = _cell_text(value, strip=False)
    if not raw:
        return [], raw
    # Match the legacy executor: trim nonblank slots, preserve all-whitespace
    # slots so that a deliberate single-space input remains executable.
    parts = [part.strip() if part.strip() else part for part in raw.split("|")]
    return parts, raw


def _priority(
    raw_value: object,
    *,
    row_number: int,
    case_code: str,
    changes: list[dict[str, object]],
) -> CasePriority:
    raw = _required_text({"priority": raw_value}, "priority")
    normalized_key = raw.upper() if raw != "中" else raw
    try:
        priority = _PRIORITY_ALIASES[normalized_key]
    except KeyError as exc:
        raise LegacyExcelMigrationError(f"不支持的优先级「{raw}」") from exc
    if raw != priority.value:
        changes.append(
            {
                "row": row_number,
                "case_code": case_code,
                "field": "priority",
                "rule": "priority_alias",
                "from": raw,
                "to": priority.value,
            }
        )
    return priority


def _enabled(raw_value: object) -> bool:
    raw = _required_text({"enabled": raw_value}, "enabled")
    normalized = raw.upper()
    if normalized in _ENABLED_VALUES:
        return True
    if normalized in _DISABLED_VALUES:
        return False
    raise LegacyExcelMigrationError(f"不支持的是否执行值「{raw}」")


def _timeout(raw_value: object) -> float | None:
    raw = _cell_text(raw_value)
    if not raw:
        return None
    try:
        value = float(raw)
    except ValueError as exc:
        raise LegacyExcelMigrationError(f"超时(秒)不是数字「{raw}」") from exc
    if not 0 < value <= 600:
        raise LegacyExcelMigrationError(f"超时(秒)超出范围「{raw}」")
    return value


def _safe_validation_message(exc: ValidationError) -> str:
    parts: list[str] = []
    for error in exc.errors(include_input=False):
        location = ".".join(str(item) for item in error["loc"])
        parts.append(f"{location}: {error['msg']}")
    return "; ".join(parts)


def _convert_case(
    row: Mapping[str, object],
    *,
    row_number: int,
    worksheet: str,
    project_key: str,
    seen_codes: set[str],
    changes: list[dict[str, object]],
    warnings: list[dict[str, object]],
) -> CaseDefinition:
    raw_case_code = _required_text(row, "用例ID")
    case_code = raw_case_code.upper()
    if case_code != raw_case_code:
        changes.append(
            {
                "row": row_number,
                "case_code": case_code,
                "field": "case_code",
                "rule": "uppercase_identifier",
                "from": raw_case_code,
                "to": case_code,
            }
        )
    if case_code in seen_codes:
        raise LegacyExcelMigrationError(f"用例ID重复「{case_code}」")
    seen_codes.add(case_code)

    module_name = _required_text(row, "模块")
    try:
        module_key = MODULE_KEYS[module_name]
    except KeyError as exc:
        raise LegacyExcelMigrationError(f"模块未配置稳定键「{module_name}」") from exc

    raw_operations = _required_text(row, "操作类型")
    operation_names = [item.strip() for item in raw_operations.split(",")]
    if any(not item for item in operation_names):
        raise LegacyExcelMigrationError("操作类型中存在空操作")

    locator_text = _cell_text(row.get("元素定位器"))
    locators = [item.strip() for item in locator_text.split(",")] if locator_text else []
    if len(locators) > len(operation_names) + 1:
        raise LegacyExcelMigrationError(
            f"定位器数量({len(locators)})超过操作数量({len(operation_names)})太多"
        )
    step_locators = locators[: len(operation_names)]
    step_locators.extend([""] * (len(operation_names) - len(step_locators)))
    extra_locator = locators[len(operation_names)] if len(locators) > len(operation_names) else ""

    data_parts, raw_input = _split_legacy_data(row.get("输入数据"))
    if row.get("输入数据") is not None and not isinstance(row.get("输入数据"), str):
        changes.append(
            {
                "row": row_number,
                "case_code": case_code,
                "field": "input_data",
                "rule": "numeric_cell_to_decimal_string",
                "from_type": type(row.get("输入数据")).__name__,
                "to_type": "string",
                "digit_count": len(raw_input or ""),
            }
        )

    steps: list[StepDefinition] = []
    input_index = 0
    for index, operation_name in enumerate(operation_names):
        spec = get_operation_spec(operation_name)
        if spec.key != operation_name:
            changes.append(
                {
                    "row": row_number,
                    "case_code": case_code,
                    "field": f"steps[{index}].operation",
                    "rule": "operation_alias",
                    "from": operation_name,
                    "to": spec.key,
                }
            )

        locator = step_locators[index] or None
        input_value: str | None = None
        if spec.consumes_input:
            input_value = data_parts[input_index] if input_index < len(data_parts) else ""
            input_index += 1
        if spec.key in {"upload", "daterange"} and not (input_value and input_value.strip()):
            raise LegacyExcelMigrationError(f"操作「{spec.key}」缺少输入数据")

        steps.append(
            StepDefinition(
                operation=spec.key,
                locator=locator,
                input=input_value,
            )
        )

    if input_index < len(data_parts):
        unused = data_parts[input_index:]
        warnings.append(
            {
                "row": row_number,
                "case_code": case_code,
                "code": "unused_input_segments",
                "message": "旧执行器不会消费这些输入槽；未写入可执行步骤",
                "segment_count": len(unused),
                "content_digest": _digest_bytes("|".join(unused).encode("utf-8")),
            }
        )

    assertion_name = _required_text(row, "断言类型")
    assertion_spec = get_assertion_spec(assertion_name)
    if assertion_spec.key != assertion_name:
        changes.append(
            {
                "row": row_number,
                "case_code": case_code,
                "field": "assertion.type",
                "rule": "assertion_alias",
                "from": assertion_name,
                "to": assertion_spec.key,
            }
        )

    explicit_assertion_locator = _cell_text(row.get("断言定位器"))
    legacy_last_locator = next((item for item in reversed(locators) if item), None)
    assertion_locator = explicit_assertion_locator or legacy_last_locator
    if explicit_assertion_locator:
        changes.append(
            {
                "row": row_number,
                "case_code": case_code,
                "field": "assertion.locator",
                "rule": "explicit_assertion_locator",
                "from": "断言定位器列",
                "to": "assertion.locator",
            }
        )
    elif extra_locator:
        changes.append(
            {
                "row": row_number,
                "case_code": case_code,
                "field": "assertion.locator",
                "rule": "promote_trailing_locator",
                "from": "元素定位器尾项",
                "to": "assertion.locator",
            }
        )
    elif assertion_spec.requires_locator and assertion_locator:
        changes.append(
            {
                "row": row_number,
                "case_code": case_code,
                "field": "assertion.locator",
                "rule": "materialize_legacy_last_locator",
                "from": "旧执行器隐式末项",
                "to": "assertion.locator",
            }
        )

    verify_point = _cell_text(row.get("验证点"))
    expected_result = _cell_text(row.get("期望结果"))
    assertion_expected = verify_point or expected_result
    assertion = AssertionDefinition(
        type=assertion_spec.key,
        expected=assertion_expected,
        locator=assertion_locator,
    )

    precondition = _cell_text(row.get("前置条件"))
    return CaseDefinition(
        case_id=uuid5(BASELINE_NAMESPACE, f"case:{project_key}:{case_code}"),
        case_code=case_code,
        module_key=module_key,
        module_name=module_name,
        title=_required_text(row, "测试场景"),
        test_point=_required_text(row, "测试点"),
        priority=_priority(
            row.get("优先级"),
            row_number=row_number,
            case_code=case_code,
            changes=changes,
        ),
        preconditions=(precondition,) if precondition else (),
        steps=tuple(steps),
        assertion=assertion,
        tags=("legacy-excel", f"module:{module_key}"),
        enabled=_enabled(row.get("是否执行")),
        source_instructions=_cell_text(row.get("操作步骤")),
        data_type=_cell_text(row.get("数据类型")),
        expected_result=expected_result,
        timeout_seconds=_timeout(row.get("超时(秒)")),
        source_trace=CaseSourceTrace(worksheet=worksheet, row_number=row_number),
        notes=_cell_text(row.get("备注")),
    )


def _counter_dict(values: Sequence[str]) -> dict[str, int]:
    return dict(sorted(Counter(values).items()))


def _build_result(
    *,
    rows: Sequence[tuple[int, Mapping[str, object]]],
    project_key: str,
    version: str,
    worksheet: str,
    source_name: str,
    source_digest: str,
    ignored_columns: Sequence[str],
) -> MigrationResult:
    changes: list[dict[str, object]] = []
    warnings: list[dict[str, object]] = []
    errors: list[dict[str, object]] = []
    seen_codes: set[str] = set()
    cases: list[CaseDefinition] = []

    for row_number, row in rows:
        try:
            case = _convert_case(
                row,
                row_number=row_number,
                worksheet=worksheet,
                project_key=project_key,
                seen_codes=seen_codes,
                changes=changes,
                warnings=warnings,
            )
        except ValidationError as exc:
            errors.append(
                {
                    "row": row_number,
                    "case_code": _cell_text(row.get("用例ID")) or "<missing>",
                    "message": _safe_validation_message(exc),
                }
            )
        except (LegacyExcelMigrationError, ValueError) as exc:
            errors.append(
                {
                    "row": row_number,
                    "case_code": _cell_text(row.get("用例ID")) or "<missing>",
                    "message": str(exc),
                }
            )
        else:
            cases.append(case)

    if errors:
        details = "\n".join(
            f"- 第{entry['row']}行[{entry['case_code']}]: {entry['message']}" for entry in errors
        )
        raise LegacyExcelMigrationError(
            f"旧 Excel 迁移失败，共 {len(errors)} 项；未生成基线：\n{details}"
        )

    baseline = CaseBaseline(
        baseline_id=uuid5(BASELINE_NAMESPACE, f"baseline:{project_key}:{version}"),
        project_key=project_key,
        version=version,
        source=CaseBaselineSource(
            file_name=source_name,
            file_digest=source_digest,
            worksheet=worksheet,
        ),
        cases=tuple(cases),
    )

    operation_names = [step.operation for case in cases for step in case.steps]
    assertion_names = [case.assertion.type for case in cases]
    priorities = [case.priority.value for case in cases]
    modules = [case.module_name for case in cases]
    enabled_count = sum(case.enabled for case in cases)
    audit: dict[str, Any] = {
        "schema_version": "1.0",
        "migration": "legacy-excel-to-case-baseline",
        "source": {
            "file_name": source_name,
            "file_digest": source_digest,
            "worksheet": worksheet,
            "ignored_columns": sorted(ignored_columns),
        },
        "target": {
            "baseline_id": str(baseline.baseline_id),
            "project_key": project_key,
            "version": version,
        },
        "policies": {
            "case_id": "UUIDv5(namespace, project_key + case_code); stable across baselines",
            "baseline_id": "UUIDv5(namespace, project_key + baseline version)",
            "input_slots": "legacy pipe order; trim nonblank slots; preserve whitespace-only slots",
            "assertion_locator": "explicit column, otherwise legacy last nonblank locator",
            "immutability": "existing artifacts may only be reused byte-for-byte",
        },
        "counts": {
            "source_cases": len(rows),
            "baseline_cases": len(cases),
            "enabled_cases": enabled_count,
            "disabled_cases": len(cases) - enabled_count,
            "changes": len(changes),
            "warnings": len(warnings),
            "errors": 0,
        },
        "distribution": {
            "modules": _counter_dict(modules),
            "priorities": _counter_dict(priorities),
            "operations": _counter_dict(operation_names),
            "assertions": _counter_dict(assertion_names),
        },
        "changes": changes,
        "warnings": warnings,
        "errors": [],
    }
    return MigrationResult(baseline=baseline, audit=audit)


def migrate_legacy_excel(
    source: str | Path,
    *,
    project_key: str,
    version: str,
    worksheet: str = DEFAULT_WORKSHEET,
) -> MigrationResult:
    """Read a legacy workbook without modifying it and build a baseline."""

    source_path = Path(source)
    if not source_path.is_file():
        raise LegacyExcelMigrationError(f"源 Excel 不存在：{source_path}")
    source_bytes = source_path.read_bytes()

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised by installation checks
        raise LegacyExcelMigrationError(
            "缺少迁移依赖 openpyxl；请安装项目的 migration 或 dev 依赖"
        ) from exc

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        if worksheet not in workbook.sheetnames:
            raise LegacyExcelMigrationError(f"Excel 中不存在工作表「{worksheet}」")
        sheet = workbook[worksheet]
        values = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(values)
        except StopIteration as exc:
            raise LegacyExcelMigrationError("Excel 工作表为空") from exc

        headers: list[str] = []
        display_headers: list[str] = []
        for index, value in enumerate(raw_headers, start=1):
            label = _cell_text(value)
            headers.append(label or f"__blank_column_{index}")
            display_headers.append(label or f"<blank-column-{index}>")
        missing = [column for column in REQUIRED_COLUMNS if column not in headers]
        if missing:
            raise LegacyExcelMigrationError(f"Excel 缺少必需列：{', '.join(missing)}")

        rows: list[tuple[int, Mapping[str, object]]] = []
        for row_number, values_row in enumerate(values, start=2):
            row = dict(zip(headers, values_row, strict=False))
            if _cell_text(row.get("用例ID")):
                rows.append((row_number, row))
        if not rows:
            raise LegacyExcelMigrationError("Excel 中没有用例")
    finally:
        workbook.close()

    ignored_columns = [
        display_headers[index]
        for index, header in enumerate(headers)
        if header not in REQUIRED_COLUMNS
    ]
    return _build_result(
        rows=rows,
        project_key=project_key,
        version=version,
        worksheet=worksheet,
        source_name=source_path.name,
        source_digest=_digest_bytes(source_bytes),
        ignored_columns=ignored_columns,
    )


def write_migration_result(result: MigrationResult, destination: str | Path) -> dict[str, Path]:
    """Write immutable artifacts; changed content requires a new version path."""

    destination_path = Path(destination)
    baseline_bytes = canonical_json_bytes(result.baseline)
    audit_bytes = canonical_json_bytes(result.audit)
    manifest = {
        "schema_version": "1.0",
        "baseline": {
            "baseline_id": str(result.baseline.baseline_id),
            "project_key": result.baseline.project_key,
            "version": result.baseline.version,
            "file": "case-baseline.json",
            "digest": _digest_bytes(baseline_bytes),
            "case_count": len(result.baseline.cases),
            "enabled_case_count": sum(case.enabled for case in result.baseline.cases),
        },
        "audit": {
            "file": "migration-audit.json",
            "digest": _digest_bytes(audit_bytes),
        },
        "source": result.baseline.source.model_dump(mode="json"),
    }
    manifest_bytes = canonical_json_bytes(manifest)
    payloads = {
        destination_path / "case-baseline.json": baseline_bytes,
        destination_path / "migration-audit.json": audit_bytes,
        destination_path / "manifest.json": manifest_bytes,
    }

    conflicts = [
        path for path, payload in payloads.items() if path.exists() and path.read_bytes() != payload
    ]
    if conflicts:
        names = ", ".join(path.name for path in conflicts)
        raise LegacyExcelMigrationError(
            f"拒绝覆盖已发布基线中的不同内容（{names}）；请创建新的 case-vX.Y.Z"
        )

    destination_path.mkdir(parents=True, exist_ok=True)
    for path, payload in payloads.items():
        if not path.exists():
            path.write_bytes(payload)
    return {
        "baseline": destination_path / "case-baseline.json",
        "audit": destination_path / "migration-audit.json",
        "manifest": destination_path / "manifest.json",
    }
